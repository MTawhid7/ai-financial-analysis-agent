"""Excel (.xlsx) exporter — structured workbook from pipeline analysis data.

One sheet per ticker + a summary sheet.
CAGR is written as a live cell formula so analysts can change the inputs.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def export_to_xlsx(
    raw_data: dict[str, Any],
    analysis: dict[str, Any],
    output_path: Path,
) -> Path:
    """Write pipeline data to an Excel workbook and save to output_path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="3B0764")  # violet
    section_fill = PatternFill(fill_type="solid", fgColor="EDE9FE")
    border = Border(
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def _header(ws, col: int, row: int, text: str) -> None:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    def _section(ws, row: int, col: int, text: str) -> None:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, size=10)
        cell.fill = section_fill

    def _val(ws, row: int, col: int, value: Any) -> None:
        ws.cell(row=row, column=col, value=value)

    # -----------------------------------------------------------------------
    # Summary sheet
    # -----------------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    _header(ws_sum, 1, 1, "Ticker")
    _header(ws_sum, 2, 1, "Sector")
    _header(ws_sum, 3, 1, "Company P/E")
    _header(ws_sum, 4, 1, "Sector P/E Avg")
    _header(ws_sum, 5, 1, "P/E Premium %")
    _header(ws_sum, 6, 1, "Price CAGR 5Y %")
    _header(ws_sum, 7, 1, "Current Price ($)")

    for i, (ticker, ta) in enumerate(analysis.items(), start=2):
        ph = (raw_data.get(ticker) or {}).get("price_history") or {}
        _val(ws_sum, i, 1, ticker)
        _val(ws_sum, i, 2, ta.get("sector", ""))
        _val(ws_sum, i, 3, ta.get("company_pe"))
        _val(ws_sum, i, 4, ta.get("sector_pe_avg"))
        _val(ws_sum, i, 5, ta.get("pe_vs_sector_premium_pct"))
        _val(ws_sum, i, 6, ta.get("price_cagr_5y_pct"))
        _val(ws_sum, i, 7, ph.get("current_price"))

    for col in range(1, 8):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    # -----------------------------------------------------------------------
    # Per-ticker sheets
    # -----------------------------------------------------------------------
    for ticker, ticker_data in raw_data.items():
        ws = wb.create_sheet(ticker[:28])  # sheet name max 31 chars
        row = 1

        # --- Price section ---
        _section(ws, row, 1, "PRICE HISTORY")
        row += 1
        ph = ticker_data.get("price_history") or {}
        for label, key in [
            ("Current Price ($)", "current_price"),
            ("Price 5 Years Ago ($)", "price_5y_ago"),
            ("52-Week High ($)", "52w_high"),
            ("52-Week Low ($)", "52w_low"),
        ]:
            _val(ws, row, 1, label)
            _val(ws, row, 2, ph.get(key))
            row += 1

        # CAGR live formula
        if ph.get("current_price") and ph.get("price_5y_ago"):
            _val(ws, row, 1, "Current Price (formula input)")
            ws.cell(row=row, column=2, value=ph["current_price"])
            cur_cell = f"B{row}"
            row += 1
            _val(ws, row, 1, "Price 5Y Ago (formula input)")
            ws.cell(row=row, column=2, value=ph["price_5y_ago"])
            ago_cell = f"B{row}"
            row += 1
            _val(ws, row, 1, "5-Year Price CAGR (live formula)")
            # Live formula: ((current/ago)^(1/5) - 1) * 100
            ws.cell(row=row, column=2, value=f"=(({cur_cell}/{ago_cell})^(1/5)-1)*100")
            ws.cell(row=row, column=2).number_format = "0.00%"

        row += 2

        # --- Fundamentals section ---
        _section(ws, row, 1, "FUNDAMENTALS")
        row += 1
        fund = ticker_data.get("fundamentals") or {}
        for label, key in [
            ("Market Cap ($)", "market_cap"),
            ("Revenue TTM ($)", "revenue_ttm"),
            ("Net Income TTM ($)", "net_income_ttm"),
            ("Profit Margin", "profit_margin"),
            ("P/E Ratio", "pe_ratio"),
            ("Forward P/E", "forward_pe"),
            ("Sector", "sector"),
            ("Industry", "industry"),
        ]:
            val = fund.get(key)
            if val is not None:
                _val(ws, row, 1, label)
                _val(ws, row, 2, val)
                row += 1

        row += 2

        # --- Balance sheet section ---
        _section(ws, row, 1, "BALANCE SHEET")
        row += 1
        bs = ticker_data.get("balance_sheet") or {}
        for label, key in [
            ("Total Assets ($)", "total_assets"),
            ("Total Liabilities ($)", "total_liabilities"),
            ("Stockholders Equity ($)", "stockholders_equity"),
            ("Cash & Equivalents ($)", "cash_and_equivalents"),
            ("Long-Term Debt ($)", "long_term_debt"),
        ]:
            val = bs.get(key)
            if val is not None:
                _val(ws, row, 1, label)
                _val(ws, row, 2, val)
                row += 1

        row += 2

        # --- Analysis section ---
        ta = analysis.get(ticker, {})
        if ta:
            _section(ws, row, 1, "QUANTITATIVE ANALYSIS")
            row += 1
            for label, key in [
                ("Sector", "sector"),
                ("Sector P/E Average", "sector_pe_avg"),
                ("Company P/E", "company_pe"),
                ("P/E vs Sector Premium (%)", "pe_vs_sector_premium_pct"),
                ("5-Year Price CAGR (%)", "price_cagr_5y_pct"),
                ("Closest Peer", "closest_peer"),
            ]:
                val = ta.get(key)
                if val is not None:
                    _val(ws, row, 1, label)
                    _val(ws, row, 2, val)
                    row += 1

            row += 1
            bull = ta.get("bull_case", [])
            bear = ta.get("bear_case", [])
            if bull:
                _section(ws, row, 1, "BULL CASE")
                row += 1
                for point in bull:
                    _val(ws, row, 1, f"• {point}")
                    row += 1
            if bear:
                row += 1
                _section(ws, row, 1, "BEAR CASE")
                row += 1
                for point in bear:
                    _val(ws, row, 1, f"• {point}")
                    row += 1

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 24

    wb.save(str(output_path))
    return output_path
