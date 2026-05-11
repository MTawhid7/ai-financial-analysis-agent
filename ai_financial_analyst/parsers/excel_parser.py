"""Excel (XLSX / XLS) file parser."""
from __future__ import annotations

import io
import re
from typing import Any

_FORMULA_RE = re.compile(r"^[=+\-@]")


def parse_xlsx(file_bytes: bytes, filename: str = "file.xlsx") -> dict[str, Any]:
    """Parse an Excel workbook. Returns per-sheet summaries (up to 10 sheets)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return {"error": f"Could not parse Excel file: {exc}", "filename": filename}

    sheets: dict[str, Any] = {}
    for name in wb.sheetnames[:10]:
        ws       = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            sheets[name] = {"rows": 0, "columns": 0}
            continue

        data_rows: list[list] = []
        row_count = injection_count = 0
        for row in rows_iter:
            row_count += 1
            clean: list[str] = []
            for cell in row:
                s = str(cell) if cell is not None else ""
                if _FORMULA_RE.match(s):
                    injection_count += 1
                    s = "[REMOVED]"
                clean.append(s[:200])
            if row_count <= 5:
                data_rows.append(clean)

        sheets[name] = {
            "rows":                  row_count,
            "columns":               len(headers),
            "column_names":          headers[:20],
            "preview":               [dict(zip(headers, r)) for r in data_rows],
            "formula_cells_removed": injection_count,
        }

    wb.close()
    return {
        "filename":    filename,
        "file_type":   "xlsx",
        "sheet_count": len(wb.sheetnames),
        "sheets":      sheets,
    }
